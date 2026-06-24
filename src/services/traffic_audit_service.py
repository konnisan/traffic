from __future__ import annotations

import csv
import io
import uuid
from datetime import UTC, datetime
from decimal import Decimal, InvalidOperation
from typing import Any

from openpyxl import load_workbook

from src.repositories.traffic_audit_repository import TrafficAuditRepository
from src.storage.postgres.manager import pg_manager
from src.storage.postgres.models_business import TrafficAuditCase
from src.utils.datetime_utils import utc_now_naive

DATASET_FIELDS = {
    "vehicles": {"record_id", "plate_number", "plate_color", "vehicle_type"},
    "transactions": {
        "record_id",
        "plate_number",
        "entry_station",
        "exit_station",
        "entry_time",
        "exit_time",
        "charged_amount",
        "expected_amount",
    },
    "passages": {
        "record_id",
        "plate_number",
        "event_time",
        "event_type",
        "location_code",
        "sequence_no",
        "next_location_code",
    },
}
REQUIRED_DATASETS = set(DATASET_FIELDS)
EVENT_TYPES = {"entry", "gantry", "exit"}


class TrafficAuditValidationError(ValueError):
    pass


def normalize_plate(value: Any) -> str:
    return "".join(str(value or "").strip().upper().split())


def parse_datetime(value: Any, field: str) -> datetime:
    if isinstance(value, datetime):
        parsed = value
    else:
        text = str(value or "").strip()
        try:
            parsed = datetime.fromisoformat(text.replace("Z", "+00:00"))
        except ValueError as exc:
            raise TrafficAuditValidationError(f"{field} 必须是 ISO 日期时间") from exc
    if parsed.tzinfo:
        return parsed.astimezone(UTC).replace(tzinfo=None)
    return parsed


def _read_rows(filename: str, content: bytes) -> list[dict[str, Any]]:
    suffix = filename.lower().rsplit(".", 1)[-1]
    if suffix == "csv":
        text = content.decode("utf-8-sig")
        return list(csv.DictReader(io.StringIO(text)))
    if suffix in {"xlsx", "xlsm"}:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        sheet = workbook.active
        values = sheet.iter_rows(values_only=True)
        try:
            headers = [str(value or "").strip() for value in next(values)]
        except StopIteration as exc:
            raise TrafficAuditValidationError("文件没有表头") from exc
        return [dict(zip(headers, row, strict=True)) for row in values]
    raise TrafficAuditValidationError("仅支持 CSV、XLSX 和 XLSM 文件")


def parse_dataset(dataset_type: str, filename: str, content: bytes) -> dict[str, Any]:
    if dataset_type not in DATASET_FIELDS:
        raise TrafficAuditValidationError(f"未知数据集类型: {dataset_type}")
    rows = _read_rows(filename, content)
    if not rows:
        raise TrafficAuditValidationError("文件没有数据行")
    missing = DATASET_FIELDS[dataset_type] - set(rows[0])
    if missing:
        raise TrafficAuditValidationError(f"缺少字段: {', '.join(sorted(missing))}")

    normalized: list[dict[str, Any]] = []
    seen: set[str] = set()
    for row_number, row in enumerate(rows, start=2):
        record_id = str(row.get("record_id") or "").strip()
        if not record_id:
            raise TrafficAuditValidationError(f"第 {row_number} 行缺少 record_id")
        if record_id in seen:
            continue
        seen.add(record_id)
        record = {key: row.get(key) for key in DATASET_FIELDS[dataset_type]}
        record["record_id"] = record_id
        record["plate_number"] = normalize_plate(record["plate_number"])
        if not record["plate_number"]:
            raise TrafficAuditValidationError(f"第 {row_number} 行缺少 plate_number")
        if dataset_type == "passages":
            record["event_type"] = str(record["event_type"] or "").strip().lower()
            if record["event_type"] not in EVENT_TYPES:
                raise TrafficAuditValidationError(f"第 {row_number} 行 event_type 必须为 entry、gantry 或 exit")
            record["event_time"] = parse_datetime(record["event_time"], "event_time").isoformat()
            record["sequence_no"] = int(record["sequence_no"])
            record["location_code"] = str(record["location_code"] or "").strip()
            record["next_location_code"] = str(record["next_location_code"] or "").strip()
        elif dataset_type == "transactions":
            record["entry_station"] = str(record["entry_station"] or "").strip()
            record["exit_station"] = str(record["exit_station"] or "").strip()
            record["entry_time"] = parse_datetime(record["entry_time"], "entry_time").isoformat()
            record["exit_time"] = parse_datetime(record["exit_time"], "exit_time").isoformat()
            for amount_field in ("charged_amount", "expected_amount"):
                try:
                    record[amount_field] = str(Decimal(str(record[amount_field])))
                except InvalidOperation as exc:
                    raise TrafficAuditValidationError(f"第 {row_number} 行 {amount_field} 不是有效金额") from exc
        record["source"] = {
            "dataset": dataset_type,
            "source_file": filename,
            "row_number": row_number,
            "record_id": record_id,
        }
        normalized.append(record)
    return {"source_file": filename, "records": normalized}


def _evidence(
    anomaly_type: str,
    severity: str,
    fact: str,
    sources: list[dict[str, Any]],
    rule_basis: str,
) -> dict[str, Any]:
    return {
        "anomaly_type": anomaly_type,
        "severity": severity,
        "fact": fact,
        "sources": sources,
        "rule_basis": rule_basis,
        "confidence_note": "由确定性规则命中，需稽核人员结合业务材料复核",
    }


def analyze_case_data(case: TrafficAuditCase) -> dict[str, Any]:
    missing = REQUIRED_DATASETS - set(case.datasets or {})
    if missing:
        raise TrafficAuditValidationError(f"尚未上传数据集: {', '.join(sorted(missing))}")
    plate = normalize_plate(case.plate_number)
    start, end = case.started_at, case.ended_at
    vehicles = [row for row in case.datasets["vehicles"]["records"] if row["plate_number"] == plate]
    if not vehicles:
        raise TrafficAuditValidationError("车辆档案中没有案件目标车辆")
    passages = [
        row
        for row in case.datasets["passages"]["records"]
        if row["plate_number"] == plate and start <= parse_datetime(row["event_time"], "event_time") <= end
    ]
    transactions = [
        row
        for row in case.datasets["transactions"]["records"]
        if row["plate_number"] == plate
        and parse_datetime(row["entry_time"], "entry_time") <= end
        and parse_datetime(row["exit_time"], "exit_time") >= start
    ]
    ordered_by_sequence = sorted(passages, key=lambda row: row["sequence_no"])
    route = sorted(passages, key=lambda row: (row["event_time"], row["sequence_no"], row["record_id"]))
    if not route:
        raise TrafficAuditValidationError("分析时间窗内没有目标车辆通行事件")
    evidence: list[dict[str, Any]] = []

    if not route or route[0]["event_type"] != "entry" or route[-1]["event_type"] != "exit":
        sources = [row["source"] for row in ([route[0], route[-1]] if route else [])]
        evidence.append(
            _evidence(
                "missing_endpoint",
                "high",
                "通行链缺少有效入口或出口事件",
                sources,
                "通行链应以入口开始并以出口结束",
            )
        )

    sequence_times = [parse_datetime(row["event_time"], "event_time") for row in ordered_by_sequence]
    if any(current < previous for previous, current in zip(sequence_times, sequence_times[1:], strict=False)):
        evidence.append(
            _evidence(
                "time_conflict",
                "high",
                "业务序号顺序与事件时间顺序不一致",
                [row["source"] for row in ordered_by_sequence],
                "同一通行链的业务序号与事件时间应保持单调递增",
            )
        )

    for current, following in zip(route, route[1:], strict=False):
        expected_next = current.get("next_location_code")
        if expected_next and expected_next != following.get("location_code"):
            evidence.append(
                _evidence(
                    "route_discontinuity",
                    "medium",
                    f"{current['location_code']} 的下一节点应为 {expected_next}，实际为 {following['location_code']}",
                    [current["source"], following["source"]],
                    "相邻通行事件应符合路段节点衔接关系",
                )
            )

    for transaction in transactions:
        charged = Decimal(transaction["charged_amount"])
        expected = Decimal(transaction["expected_amount"])
        if charged != expected:
            evidence.append(
                _evidence(
                    "fee_mismatch",
                    "high",
                    f"实收 {charged} 元，规则应收 {expected} 元，差额 {expected - charged} 元",
                    [transaction["source"]],
                    "收费交易实收金额应与规则计算金额一致",
                )
            )

    return {
        "case_id": case.id,
        "plate_number": plate,
        "route": route,
        "evidence": evidence,
        "risk_summary": "发现疑似异常，等待人工复核" if evidence else "未发现规则异常",
        "analyzed_at": utc_now_naive().isoformat(),
    }


def build_report(case: TrafficAuditCase, result: dict[str, Any]) -> str:
    lines = [
        f"# {case.title}",
        "",
        "## 案件概况",
        f"- 车牌号：{case.plate_number}",
        f"- 分析时间窗：{case.started_at.isoformat()} 至 {case.ended_at.isoformat()}",
        f"- 辅助研判结论：{result['risk_summary']}",
        "- 正式结论：未经人工确认",
        "",
        "## 路径还原",
    ]
    lines.extend(
        f"- {row['event_time']} | {row['event_type']} | {row['location_code']} | "
        f"来源 {row['source']['source_file']}:{row['source']['row_number']}"
        for row in result["route"]
    )
    lines.extend(["", "## 异常证据"])
    if not result["evidence"]:
        lines.append("- 未命中一期规则。")
    for index, item in enumerate(result["evidence"], start=1):
        references = ", ".join(
            f"{src['source_file']}:{src['row_number']}({src['record_id']})" for src in item["sources"]
        )
        lines.extend(
            [
                f"### {index}. {item['anomaly_type']}",
                f"- 风险等级：{item['severity']}",
                f"- 事实：{item['fact']}",
                f"- 规则依据：{item['rule_basis']}",
                f"- 源记录：{references or '无匹配记录'}",
                f"- 置信说明：{item['confidence_note']}",
            ]
        )
    lines.extend(["", "## 人工复核", "本报告为智能体生成的辅助研判草稿，必须由稽核人员复核后使用。"])
    return "\n".join(lines)


async def run_case_analysis(case_id: str) -> dict[str, Any]:
    async with pg_manager.get_async_session_context() as db:
        repository = TrafficAuditRepository(db)
        case = await repository.get(case_id)
        if not case:
            raise TrafficAuditValidationError("案件不存在")
        case.status = "analyzing"
        await repository.save(case)
        try:
            result = analyze_case_data(case)
            case.analysis_result = result
            case.report_markdown = build_report(case, result)
            case.status = "pending_review"
            await repository.save(case)
            return result
        except Exception:
            case.status = "analysis_failed"
            await repository.save(case)
            await db.commit()
            raise


def new_case(
    *, title: str, plate_number: str, started_at: datetime, ended_at: datetime, department_id: int, created_by: str
) -> TrafficAuditCase:
    normalized_start = parse_datetime(started_at, "started_at")
    normalized_end = parse_datetime(ended_at, "ended_at")
    if normalized_end <= normalized_start:
        raise TrafficAuditValidationError("结束时间必须晚于开始时间")
    plate = normalize_plate(plate_number)
    if not plate:
        raise TrafficAuditValidationError("车牌号不能为空")
    if not title.strip():
        raise TrafficAuditValidationError("案件名称不能为空")
    return TrafficAuditCase(
        id=uuid.uuid4().hex,
        title=title.strip(),
        plate_number=plate,
        started_at=normalized_start,
        ended_at=normalized_end,
        department_id=department_id,
        created_by=created_by,
        datasets={},
    )
