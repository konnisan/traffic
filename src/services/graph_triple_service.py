from __future__ import annotations

import csv
import io
import json
import uuid
from typing import Any

from src.repositories.graph_triple_repository import GraphTripleRepository

ALLOWED_SOURCE_KINDS = {"curated_triples", "simple_extracted_relations"}
ALLOWED_STATUSES = {"draft", "approved", "rejected"}


class GraphTripleService:
    def __init__(self, repo: GraphTripleRepository | None = None):
        self.repo = repo or GraphTripleRepository()

    async def import_file(self, filename: str, content: bytes, created_by: str) -> dict[str, Any]:
        triples = self._parse_file(filename, content)
        rows = []
        for item in triples:
            normalized = self._normalize(item)
            normalized["triple_id"] = f"gtd_{uuid.uuid4().hex[:16]}"
            normalized["status"] = "draft"
            normalized["created_by"] = created_by
            rows.append(normalized)

        created = await self.repo.create_many(rows)
        return {"imported": len(created), "triples": [self._to_dict(row) for row in created]}

    async def list_triples(
        self, status: str | None = None, source_kind: str | None = None, limit: int = 100
    ) -> list[dict[str, Any]]:
        if status and status not in ALLOWED_STATUSES:
            raise ValueError(f"Invalid status: {status}")
        if source_kind and source_kind not in ALLOWED_SOURCE_KINDS:
            raise ValueError(f"Invalid source_kind: {source_kind}")
        rows = await self.repo.list(status=status, source_kind=source_kind, limit=limit)
        return [self._to_dict(row) for row in rows]

    async def review_triple(
        self, triple_id: str, action: str, reviewed_by: str, review_comment: str | None = None
    ) -> dict[str, Any]:
        if action not in {"approve", "reject"}:
            raise ValueError("action must be approve or reject")
        target_status = "approved" if action == "approve" else "rejected"
        row = await self.repo.update_status(triple_id, target_status, reviewed_by, review_comment)
        if not row:
            raise ValueError(f"Triple not found: {triple_id}")
        return self._to_dict(row)

    async def get_triple(self, triple_id: str) -> dict[str, Any]:
        row = await self.repo.get(triple_id)
        if not row:
            raise ValueError(f"Triple not found: {triple_id}")
        return self._to_dict(row)

    @staticmethod
    def to_graph_payload(item: dict[str, Any]) -> dict[str, Any]:
        return {
            "h": {
                "name": item["subject"],
                "type": item.get("subject_type") or "unknown",
                "source_kind": item.get("source_kind"),
                "source_doc_id": item.get("source_doc_id"),
            },
            "r": {
                "type": item["predicate"],
                "source_doc_id": item.get("source_doc_id"),
                "source_page": item.get("source_page"),
                "source_quote": item.get("source_quote"),
                "confidence": item.get("confidence"),
                "review_status": item.get("status"),
            },
            "t": {
                "name": item["object"],
                "type": item.get("object_type") or "unknown",
                "source_kind": item.get("source_kind"),
                "source_doc_id": item.get("source_doc_id"),
            },
        }

    def _parse_file(self, filename: str, content: bytes) -> list[dict[str, Any]]:
        suffix = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
        if suffix == "jsonl":
            return [json.loads(line) for line in content.decode("utf-8-sig").splitlines() if line.strip()]
        if suffix == "csv":
            text = content.decode("utf-8-sig")
            return list(csv.DictReader(io.StringIO(text)))
        if suffix in {"xlsx", "xlsm"}:
            from openpyxl import load_workbook

            workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
            sheet = workbook.active
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                return []
            headers = [str(cell or "").strip() for cell in rows[0]]
            return [
                {headers[index]: value for index, value in enumerate(row) if index < len(headers)}
                for row in rows[1:]
                if any(value is not None and str(value).strip() for value in row)
            ]
        raise ValueError("Only jsonl, csv, xlsx and xlsm triple files are supported")

    def _normalize(self, item: dict[str, Any]) -> dict[str, Any]:
        if "h" in item or "r" in item or "t" in item:
            h = item.get("h") or {}
            r = item.get("r") or {}
            t = item.get("t") or {}
            item = {
                "subject": h.get("name") if isinstance(h, dict) else h,
                "predicate": r.get("type") if isinstance(r, dict) else r,
                "object": t.get("name") if isinstance(t, dict) else t,
                "subject_type": h.get("type") if isinstance(h, dict) else None,
                "object_type": t.get("type") if isinstance(t, dict) else None,
                "source_doc_id": r.get("source_doc_id") if isinstance(r, dict) else None,
                "source_page": r.get("source_page") if isinstance(r, dict) else None,
                "source_quote": r.get("source_quote") if isinstance(r, dict) else None,
                "confidence": r.get("confidence") if isinstance(r, dict) else None,
                "source_kind": r.get("source_kind") if isinstance(r, dict) else None,
            }

        subject = str(item.get("subject") or "").strip()
        predicate = str(item.get("predicate") or "").strip()
        obj = str(item.get("object") or "").strip()
        if not subject or not predicate or not obj:
            raise ValueError("subject, predicate and object are required")

        source_kind = str(item.get("source_kind") or "curated_triples").strip()
        if source_kind not in ALLOWED_SOURCE_KINDS:
            raise ValueError(f"Invalid source_kind: {source_kind}")

        confidence = item.get("confidence")
        confidence_value = float(confidence) if confidence not in (None, "") else 1.0
        confidence_value = max(0.0, min(confidence_value, 1.0))

        source_page = item.get("source_page")
        source_page_value = int(source_page) if source_page not in (None, "") else None

        return {
            "subject": subject,
            "predicate": predicate,
            "object": obj,
            "subject_type": str(item.get("subject_type") or "").strip() or None,
            "object_type": str(item.get("object_type") or "").strip() or None,
            "source_doc_id": str(item.get("source_doc_id") or "").strip() or None,
            "source_page": source_page_value,
            "source_quote": str(item.get("source_quote") or "").strip() or None,
            "confidence": confidence_value,
            "source_kind": source_kind,
        }

    @staticmethod
    def _to_dict(row) -> dict[str, Any]:
        return {
            "triple_id": row.triple_id,
            "subject": row.subject,
            "predicate": row.predicate,
            "object": row.object,
            "subject_type": row.subject_type,
            "object_type": row.object_type,
            "source_doc_id": row.source_doc_id,
            "source_page": row.source_page,
            "source_quote": row.source_quote,
            "confidence": row.confidence,
            "source_kind": row.source_kind,
            "status": row.status,
            "review_comment": row.review_comment,
            "created_by": row.created_by,
            "reviewed_by": row.reviewed_by,
            "approved_at": row.approved_at.isoformat() if row.approved_at else None,
            "created_at": row.created_at.isoformat() if row.created_at else None,
            "updated_at": row.updated_at.isoformat() if row.updated_at else None,
        }
